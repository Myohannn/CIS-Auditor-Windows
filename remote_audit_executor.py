import pandas as pd
import time
import logging
import argparse
import sys


from multiprocessing import Pool, Manager, freeze_support
from openpyxl import load_workbook

from utilities.getPwdPolicy import get_pwd_policy_actual_value, compare_pwd_policy
from utilities.getRegValue import get_registry_actual_value, compare_reg_value
from utilities.getLockoutPolicy import get_lockout_policy_actual_value, compare_lockout_policy
from utilities.getUserRights import get_user_rights_actual_value, compare_user_rights
from utilities.getCheckAccount import get_check_account_actual_value, compare_check_account
from utilities.getBannerCheck import get_banner_check_actual_value, compare_banner_check
from utilities.getAnonySID import get_anonymous_sid_value, compare_anonymous_sid
from utilities.getAuditPolicy import get_audit_policy_actual_value, compare_audit_policy
from utilities.getRegCheck import get_reg_check_actual_value, compare_reg_check
from utilities.getWMIPolicy import get_wmi_policy_actual_value, compare_wmi_policy


def gen_ps_args(data_dict: dict) -> dict:
    '''
    This function generates a dictionary of PowerShell commands for each audit type
    based on the audit data provided.

    :param data_dict: 
        A dictionary with keys representing different audit types and values as 
        DataFrames containing the audit data.
    :return: 
        A dictionary with the same keys as data_dict but the values replaced 
        with lists of PowerShell commands for each audit type.
    '''

    ps_args_dict = {}

    for key, df in data_dict.items():

        checklist_values = df['Checklist'].values
        description_values = df['Description'].values
        reg_key_values = df['Reg Key'].values
        reg_item_values = df['Reg Item'].values
        subcategory_values = df['Audit Policy Subcategory'].values
        right_type_values = df['Right type'].values

        actual_value_list = []

        if key == "REGISTRY_SETTING":

            reg_value_args = []
            reg_value_args_list = []

            for idx, val in enumerate(checklist_values):
                # # if val == 1:
                # generate command list for getting regristry value
                reg_key = str(reg_key_values[idx])
                reg_item = str(reg_item_values[idx])

                if reg_key.startswith("HKLM"):
                    reg_key = reg_key.replace("HKLM", "HKLM:")
                elif reg_key.startswith("HKU"):
                    reg_key = reg_key.replace("HKU", "HKU:")

                arg = f"Write-Output '====';Get-ItemPropertyValue -Path '{reg_key}' -Name '{reg_item}'"
                reg_value_args.append(arg)

                if len(reg_value_args) == 50:
                    reg_value_args_list.append(
                        ';'.join(reg_value_args))
                    reg_value_args = []

            reg_value_args_list.append(';'.join(reg_value_args))

            ps_args_dict[key] = reg_value_args_list

            continue

        elif key == "PASSWORD_POLICY":

            pwd_policy_args = []
            pwd_policy_args_list = []

            for idx, val in enumerate(checklist_values):
                # if val == 1:

                # generate command list for getting password policy value
                description = str(description_values[idx])

                if "Enforce password history" in description:
                    subcategory = 'PasswordHistorySize ='

                elif "Maximum password age" in description:
                    subcategory = 'MaximumPasswordAge ='

                elif "Minimum password age" in description:
                    subcategory = 'MinimumPasswordAge ='

                elif "Minimum password length" in description:
                    subcategory = 'MinimumPasswordLength ='

                elif "complexity requirements" in description:
                    subcategory = 'PasswordComplexity ='

                elif "reversible encryption" in description:
                    subcategory = 'ClearTextPassword ='

                elif "Administrator account lockout" in description:
                    subcategory = ''

                elif "Force logoff when logon hours expire" in description:
                    subcategory = 'ForceLogoffWhenHourExpire ='

                else:
                    actual_value_list.append("")

                arg = f"Write-Output '===='; Get-Content -Path C:\\temp\\secpol.cfg | Select-String -Pattern '{subcategory}'"

                pwd_policy_args.append(arg)

                if len(pwd_policy_args) == 50:
                    pwd_policy_args_list.append(
                        ';'.join(pwd_policy_args))
                    pwd_policy_args = []

            pwd_policy_args_list.append(';'.join(pwd_policy_args))

            ps_args_dict[key] = pwd_policy_args_list

            continue

        elif key == "LOCKOUT_POLICY":

            lockout_policy_args = []
            for idx, val in enumerate(checklist_values):
                # if val == 1:

                # generate command list for getting lockout policy value
                description = str(description_values[idx])

                if "Account lockout duration" in description:
                    lockout_policy_args.append(
                        "net accounts | select-string -pattern 'Lockout duration'")

                elif "Account lockout threshold" in description:
                    lockout_policy_args.append(
                        "net accounts | select-string -pattern 'Lockout threshold'")

                elif "Reset account lockout counter" in description:
                    lockout_policy_args.append(
                        "net accounts | select-string -pattern 'Lockout observation window'")
                else:
                    lockout_policy_args.append("")

            ps_args_dict[key] = lockout_policy_args

        elif key == "USER_RIGHTS_POLICY":

            user_rights_args = []
            user_rights_args_list = []

            for idx, val in enumerate(checklist_values):

                # if val == 1:
                right_type = str(right_type_values[idx])

                arg = f"Write-Output '===='; Get-Content -Path C:\\temp\\secpol.cfg | Select-String -Pattern '{right_type}'"

                user_rights_args.append(arg)

                if len(user_rights_args) == 50:
                    user_rights_args_list.append(
                        ';'.join(user_rights_args))
                    user_rights_args = []

            user_rights_args_list.append(';'.join(user_rights_args))

            ps_args_dict[key] = user_rights_args_list

        elif key == "CHECK_ACCOUNT":
            check_account_args = []
            for idx, val in enumerate(checklist_values):
                # if val == 1:

                # generate command list for getting check account value
                description = str(description_values[idx])

                if "Guest account status" in description:
                    check_account_args.append(
                        "net user guest | select-string -pattern 'Account active'")

                elif "Administrator account status" in description:
                    check_account_args.append(
                        "net user administrator | select-string -pattern 'Account active'")

                elif "Rename administrator account" in description:
                    check_account_args.append(
                        "net user administrator | select-string -pattern 'User name'")

                elif "Rename guest account" in description:
                    check_account_args.append(
                        "net user guest | select-string -pattern 'User name'")
                else:
                    check_account_args.append("")

            ps_args_dict[key] = check_account_args

        elif key == "BANNER_CHECK":
            banner_check_args = []
            banner_check_args_list = []

            for idx, val in enumerate(checklist_values):
                # if val == 1:
                # generate command list for getting regristry value
                reg_key = str(reg_key_values[idx])
                reg_item = str(reg_item_values[idx])

                if reg_key.startswith("HKLM"):
                    reg_key = reg_key.replace("HKLM", "HKLM:")
                elif reg_key.startswith("HKU"):
                    reg_key = reg_key.replace("HKU", "HKU:")

                arg = f"Write-Output '===='; Get-ItemPropertyValue -Path '{reg_key}' -Name '{reg_item}'"
                banner_check_args.append(arg)

                if len(banner_check_args) == 50:
                    banner_check_args_list.append(
                        ';'.join(banner_check_args))
                    banner_check_args = []

            banner_check_args_list.append(';'.join(banner_check_args))

            ps_args_dict[key] = banner_check_args_list

        elif key == "ANONYMOUS_SID_SETTING":

            anonymous_sid_args = []
            anonymous_sid_args_list = []

            for idx, val in enumerate(checklist_values):
                # if val == 1:

                # generate command list for getting password policy value
                description = str(description_values[idx])

                if "Allow anonymous SID/Name translation" in description:
                    subcategory = 'LSAAnonymousNameLookup ='
                else:
                    actual_value_list.append("")

                arg = f"Write-Output '===='; Get-Content -Path C:\\temp\\secpol.cfg | Select-String -Pattern '{subcategory}'"

                anonymous_sid_args.append(arg)

                if len(anonymous_sid_args) == 50:
                    anonymous_sid_args_list.append(
                        ';'.join(anonymous_sid_args))
                    anonymous_sid_args = []

            anonymous_sid_args_list.append(';'.join(anonymous_sid_args))

            ps_args_dict[key] = anonymous_sid_args_list

            continue

        elif key == "AUDIT_POLICY_SUBCATEGORY":
            audit_policy_args = []
            audit_policy_args_list = []

            for idx, val in enumerate(checklist_values):

                # if val == 1:
                subcategory = str(subcategory_values[idx])

                arg = f"Write-Output '===='; auditpol /get /subcategory:'{subcategory}' | select-string -pattern '{subcategory}'"

                audit_policy_args.append(arg)

                if len(audit_policy_args) == 50:
                    audit_policy_args_list.append(
                        ';'.join(audit_policy_args))
                    audit_policy_args = []

            audit_policy_args_list.append(';'.join(audit_policy_args))

            ps_args_dict[key] = audit_policy_args_list

        elif key == "REG_CHECK":
            reg_check_args = []
            reg_check_args_list = []

            for idx, val in enumerate(checklist_values):

                # if val == 1:

                reg_key = reg_key_values[idx]
                reg_item = reg_item_values[idx]

                if reg_key.startswith("HKLM"):
                    reg_key = reg_key.replace("HKLM", "HKLM:")
                elif reg_key.startswith("HKU"):
                    reg_key = reg_key.replace("HKU", "HKU:")

                arg = f"Write-Output '====';Get-ItemPropertyValue -Path '{reg_key}' -Name '{reg_item}'"

                reg_check_args.append(arg)

            reg_check_args_list.append(';'.join(reg_check_args))

            ps_args_dict[key] = reg_check_args_list

        elif key == "WMI_POLICY":
            wmi_policy_args = []
            wmi_policy_args_list = []

            for idx, val in enumerate(checklist_values):

                # if val == 1:

                arg = "Write-Output '====';(Get-WmiObject -Class Win32_ComputerSystem).DomainRole"

                wmi_policy_args.append(arg)

            wmi_policy_args_list.append(';'.join(wmi_policy_args))

            ps_args_dict[key] = wmi_policy_args_list

        else:
            continue

    return ps_args_dict


def get_actual_values(ip: list, ps_args_dict: dict, data_dict: dict) -> dict:
    '''
    This function executes the PowerShell commands generated by `gen_ps_args()`
    on the target system and retrieves their outputs. It then compares the actual 
    and expected values for each audit type and returns a new dictionary with 
    the comparison results.

    :param ip: 
        A list containing the IP address, username, password, and Windows version 
        of the target system.
    :param ps_args_dict: 
        A dictionary with keys representing different audit types and values as 
        lists of PowerShell commands.
    :param data_dict: 
        A dictionary with keys representing different audit types and values as 
        DataFrames containing the audit data.
    :return: 
        A new dictionary with the same keys as data_dict but the values replaced 
        with DataFrames that include the results of the comparison between the 
        actual and expected values.
    '''

    new_dict = {}

    for key, args_list in ps_args_dict.items():

        try:

            if key == "PASSWORD_POLICY":
                logging.info(f"{ip[0]} | Getting {key} value......")
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_pwd_policy_actual_value(
                        args_list, ip)

                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_pwd_policy(
                    ip[0], actual_value_list, data_dict)
            elif key == "REGISTRY_SETTING":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_registry_actual_value(
                        args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_reg_value(ip[0], actual_value_list, data_dict)
            elif key == "LOCKOUT_POLICY":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_lockout_policy_actual_value(
                        args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_lockout_policy(
                    ip[0], actual_value_list, data_dict)
            elif key == "USER_RIGHTS_POLICY":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_user_rights_actual_value(
                        args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_user_rights(
                    ip[0], actual_value_list, data_dict)
            elif key == "CHECK_ACCOUNT":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_check_account_actual_value(
                        args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_check_account(
                    ip[0], actual_value_list, data_dict)
            elif key == "BANNER_CHECK":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                actual_value_list = get_banner_check_actual_value(
                    args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_banner_check(
                    ip[0], actual_value_list, data_dict)
            elif key == "ANONYMOUS_SID_SETTING":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_anonymous_sid_value(args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_anonymous_sid(
                    ip[0], actual_value_list, data_dict)
            elif key == "AUDIT_POLICY_SUBCATEGORY":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_audit_policy_actual_value(
                        args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_audit_policy(
                    ip[0], actual_value_list, data_dict)
            elif key == "REG_CHECK":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_reg_check_actual_value(
                        args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_reg_check(ip[0], actual_value_list, data_dict)
            elif key == "WMI_POLICY":
                logging.info(f"{ip[0]} | Getting {key} value......")
                # continue
                if args_list == [''] or args_list == []:
                    actual_value_list = []
                else:
                    actual_value_list = get_wmi_policy_actual_value(
                        args_list, ip)
                logging.info(f"{ip[0]} | Comparing {key} value......")
                new_df = compare_wmi_policy(
                    ip[0], actual_value_list, data_dict)

            new_dict[key] = new_df
        except Exception as e:
            logging.debug('Failed to get actual value: %s', e)

        # new_dict[key] = new_df

    return new_dict


def read_file(win_version: str) -> dict:
    '''
    This function reads the provided audit file based on the Windows version 
    and returns a dictionary where the keys are different audit types and the 
    values are corresponding audit data in DataFrame format.

    :param win_version: 
        A string representing the Windows version of the target system.
    :return: 
        A dictionary with keys representing different audit types and values as 
        DataFrames containing the audit data.
    '''

    data_dict = {
        "PASSWORD_POLICY": [],
        "REGISTRY_SETTING": [],
        "LOCKOUT_POLICY": [],
        "USER_RIGHTS_POLICY": [],
        "CHECK_ACCOUNT": [],
        "BANNER_CHECK": [],
        "ANONYMOUS_SID_SETTING": [],
        "AUDIT_POLICY_SUBCATEGORY": [],
        "REG_CHECK": [],
        "WMI_POLICY": []
    }

    version_dict = {
        'Windows 10 Enterprise': 'src\Audit\CIS_MS_Windows_10_Enterprise_Level_1_v2.0.0.xlsx',
        'Windows 11 Enterprise': 'src\Audit\CIS_MS_Windows_11_Enterprise_Level_1_v1.0.0.xlsx',
        'Windows Server 2016 MS': 'src\Audit\CIS_Microsoft_Windows_Server_2016_Benchmark_v2.0.0_L1_MS.xlsx',
        'Windows Server 2019 MS': 'src\Audit\CIS_Microsoft_Windows_Server_2019_Benchmark_v2.0.0_L1_MS.xlsx',
        'Windows Server 2019 DC': 'src\Audit\CIS_Microsoft_Windows_Server_2019_Benchmark_v2.0.0_L1_DC.xlsx',
        'Windows Server 2022 MS': 'src\Audit\CIS_Microsoft_Windows_Server_2022_Benchmark_v2.0.0_L1_MS.xlsx',

    }

    fname = version_dict[win_version]

    xl = pd.ExcelFile(fname)
    # df = xl.parse(sheet_name=0)

    for type in data_dict:
        try:
            data_dict[type] = xl.parse(sheet_name=type)
        except ValueError as e:
            logging.error(f"{type} not found")

    return data_dict


def save_file(out_fname: str, data_dict_list: list) -> None:
    '''
    This function processes the comparison results and saves them into an Excel file.

    :param out_fname: 
        A string representing the filename of the output file.
    :param data_dict_list: 
        A list of dictionaries containing comparison results.
    :return: 
        None
    '''

    if data_dict_list == []:
        return

    all_frames = []

    for data_dict in data_dict_list:
        frames = []

        for df in data_dict.values():
            frames.append(df)
        result_rowwise = pd.concat(frames)
        all_frames.append(result_rowwise.reset_index(drop=True))

    result = pd.concat(all_frames, axis=1)

    result = result.loc[:, ~result.columns.duplicated()]

    column_names = result.columns.tolist()

    value_n_result = column_names[11:]
    ip_list = []
    name_list = []

    for i in range(len(value_n_result)):
        if i % 2 == 0:
            ip = value_n_result[i].split('|')[0].strip()
            ip_list.append(ip)
            name_list.append('Actual Value')
        else:
            ip_list.append('')
            name_list.append('Result')

    new_data = ['Checklist', 'Type', 'Index', 'Description', 'Solution', 'Reg Key', 'Reg Item', 'Reg Option', 'Audit Policy Subcategory',
                'Right type', 'Value Data'] + name_list
    result.columns = ['Checklist', 'Type', 'Index', 'Description', 'Solution', 'Reg Key', 'Reg Item', 'Reg Option', 'Audit Policy Subcategory',
                      'Right type', 'Value Data'] + ip_list

    new_df = pd.DataFrame(
        [new_data + [''] * (result.shape[1] - len(new_data))], columns=result.columns)
    result = pd.concat([new_df, result]).reset_index(drop=True)

    # Save DataFrame to a new Excel file
    result.to_excel(out_fname, index=False)

    # Load the workbook and select the sheet
    wb = load_workbook(out_fname)
    ws = wb.active

    # Merge the appropriate cells in the new first row
    for col in range(1, 12):  # adjust these values as needed
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=2, end_column=col)

    for ip_col in range(12, len(result.columns)):  # adjust these values as needed
        if ip_col % 2 == 0:
            ws.merge_cells(start_row=1, start_column=ip_col,
                           end_row=1, end_column=ip_col+1)
        else:
            continue

    # Save the workbook
    wb.save(out_fname)
    print((f"Result saved into {out_fname}"))

    logging.info(f"Result saved into {out_fname}")


def configurations(config_fname: str) -> list:
    '''
    This function reads the configuration settings from the 'config.xlsx' file,
    including IP address, user account, and Windows version.

    :param config_fname: 
        A string representing the filename of the configuration file.
    :return: 
        A list of lists where each sub-list contains the IP address, username, 
        password, and Windows version of a target system.
    '''

    config_file = pd.ExcelFile(config_fname)

    configs = config_file.parse(sheet_name=0)

    ip_list = []

    ips = configs['IP Address'].values
    usernames = configs['Username'].values
    passwords = configs['Password'].values
    versions = configs['Windows Version'].values

    for idx, val in enumerate(ips):
        ip = [ips[idx], usernames[idx],
              passwords[idx], versions[idx]]

        ip_list.append(ip)

    return ip_list


def run(ip: str, data_dict: dict) -> dict:
    '''
    This function runs the audit process for a single IP address.

    :param ip: 
        A string representing the IP address of the target system.
    :param data_dict: 
        A dictionary with keys representing different audit types and values as 
        DataFrames containing the audit data.
    :return: 
        A new dictionary with the same keys as data_dict but the values replaced 
        with DataFrames that include the results of the comparison between the 
        actual and expected values.
    '''

    logging.info(f'IP: {ip[0]} scanning start....')
    start_t = time.time()

    try:
        ps_args_dict = gen_ps_args(data_dict)

    except Exception as e:
        logging.error('Failed to generate powershell command')
        exit()

    # get actual values and compare
    new_dict = get_actual_values(ip, ps_args_dict, data_dict)

    end_t = time.time()

    running_t = end_t - start_t
    logging.info(f"IP: {ip[0]} | Running time: {running_t} seconds")
    return new_dict


'''
This is the main function that gets executed when the script runs. It parses 
command-line arguments, reads the configuration file, reads the audit file, 
initiates parallel processing for multiple IP addresses, and finally saves 
the audit results into an output file.
'''
if __name__ == '__main__':

    freeze_support()

    my_parser = argparse.ArgumentParser(
        description='This is a script for performing a customizable, multiprocessing security audit on multiple remote systems.')

    # Add the arguments
    my_parser.add_argument(
        '-config',
        type=str,
        required=True,
        help='(REQUIRED) The path to the configuration file that contains the IP addresses, usernames, passwords, and Windows versions of the systems to be audited. This should be a .xlsx file'
    )

    my_parser.add_argument(
        '-output',
        type=str,
        required=True,
        help='(REQUIRED) The path to the output file where the results of the audit will be saved. This should be a .xlsx file.'
    )

    # Execute parse_args()
    try:
        args = my_parser.parse_args()
    except SystemExit:
        my_parser.print_help()
        sys.exit(1)

    print('Configuration file:', args.config)
    print('Output file:', args.output)

    start_t0 = time.time()

    # read configurations
    ip_list = configurations(args.config)

    for i in ip_list:
        print(f"IP: {i[0]} - {i[3]} loaded")

    win_version = ip_list[0][3]

    data_dict = read_file(win_version)

    # Initiate multiprocess
    with Manager() as manager:
        # initialize shared dictionary with data_dict
        shared_data_dict = manager.dict(data_dict)

        with Pool(processes=4) as pool:
            results = pool.starmap(
                run, [(ip, shared_data_dict) for ip in ip_list])

    # write output file
    save_file(args.output, results)

    start_t = time.time()

    end_t0 = time.time()

    running_t0 = end_t0 - start_t0
    logging.info(f"The script ran for {running_t0} seconds")
