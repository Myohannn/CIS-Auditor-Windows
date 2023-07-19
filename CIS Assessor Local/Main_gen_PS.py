import pandas as pd
import logging
import argparse

from openpyxl import load_workbook

# set up logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

handler = logging.FileHandler('mylog.log', mode='w')
handler.setLevel(logging.DEBUG)

formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

logger.addHandler(handler)


def gen_ps_args(data_dict):
    ps_args_dict = {}

    for key, df in data_dict.items():

        checklist_values = df['Checklist'].values
        description_values = df['Description'].values
        reg_key_values = df['Reg Key'].values
        reg_item_values = df['Reg Item'].values
        subcategory_values = df['Audit Policy Subcategory'].values
        right_type_values = df['Right type'].values

        if key == "REGISTRY_SETTING":

            reg_value_args = []
            reg_value_args_list = []

            for idx, val in enumerate(checklist_values):
                # # if val == 1:
                # generate command list for getting regristry value
                reg_key = str(reg_key_values[idx])
                reg_item = str(reg_item_values[idx])

                if reg_key.startswith("HKLM"):
                    reg_key = reg_key.replace("HKLM", "HKLM:")
                elif reg_key.startswith("HKU"):
                    reg_key = reg_key.replace("HKU", "HKU:")

                arg = f"Write-Output '====';Get-ItemPropertyValue -Path '{reg_key}' -Name '{reg_item}'"
                reg_value_args.append(arg)

                # if len(reg_value_args) == 50:
                #     reg_value_args_list.append(
                #         ';'.join(reg_value_args))
                #     reg_value_args = []

            reg_value_args_list.append(';'.join(reg_value_args))

            ps_args_dict[key] = ';'.join(reg_value_args)

            continue

        elif key == "PASSWORD_POLICY":

            pwd_policy_args = []
            pwd_policy_args_list = []

            for idx, val in enumerate(checklist_values):
                # if val == 1:

                # generate command list for getting password policy value
                description = str(description_values[idx])

                if "Enforce password history" in description:
                    subcategory = 'PasswordHistorySize ='

                elif "Maximum password age" in description:
                    subcategory = 'MaximumPasswordAge ='

                elif "Minimum password age" in description:
                    subcategory = 'MinimumPasswordAge ='

                elif "Minimum password length" in description:
                    subcategory = 'MinimumPasswordLength ='

                elif "complexity requirements" in description:
                    subcategory = 'PasswordComplexity ='

                elif "reversible encryption" in description:
                    subcategory = 'ClearTextPassword ='

                elif "Administrator account lockout" in description:
                    subcategory = ''

                elif "Force logoff when logon hours expire" in description:
                    subcategory = 'ForceLogoffWhenHourExpire ='

                else:
                    pwd_policy_args.append("")

                arg = f"Write-Output '===='; Get-Content -Path C:\\temp\\secpol.cfg | Select-String -Pattern '{subcategory}'"

                pwd_policy_args.append(arg)

                # if len(pwd_policy_args) == 50:
                #     pwd_policy_args_list.append(
                #         ';'.join(pwd_policy_args))
                #     pwd_policy_args = []

            pwd_policy_args_list.append(';'.join(pwd_policy_args))

            ps_args_dict[key] = ';'.join(pwd_policy_args)

            continue

        elif key == "LOCKOUT_POLICY":

            lockout_policy_args = []
            for idx, val in enumerate(checklist_values):
                # if val == 1:

                # generate command list for getting lockout policy value
                description = str(description_values[idx])

                if "Account lockout duration" in description:
                    lockout_policy_args.append(
                        "Write-Output '====';net accounts | select-string -pattern 'Lockout duration'")

                elif "Account lockout threshold" in description:
                    lockout_policy_args.append(
                        "Write-Output '====';net accounts | select-string -pattern 'Lockout threshold'")

                elif "Reset account lockout counter" in description:
                    lockout_policy_args.append(
                        "Write-Output '====';net accounts | select-string -pattern 'Lockout observation window'")
                else:
                    lockout_policy_args.append("Write-Output '====';")

            ps_args_dict[key] = ';'.join(lockout_policy_args)

        elif key == "USER_RIGHTS_POLICY":

            user_rights_args = []
            user_rights_args_list = []

            for idx, val in enumerate(checklist_values):

                # if val == 1:
                right_type = str(right_type_values[idx])

                arg = f"Write-Output '===='; Get-Content -Path C:\\temp\\secpol.cfg | Select-String -Pattern '{right_type}'"

                user_rights_args.append(arg)

                # if len(user_rights_args) == 50:
                #     user_rights_args_list.append(
                #         ';'.join(user_rights_args))
                #     user_rights_args = []

            user_rights_args_list.append(';'.join(user_rights_args))

            ps_args_dict[key] = ';'.join(user_rights_args)

        elif key == "CHECK_ACCOUNT":
            check_account_args = []
            for idx, val in enumerate(checklist_values):
                # if val == 1:

                # generate command list for getting check account value
                description = str(description_values[idx])

                if "Guest account status" in description:
                    check_account_args.append(
                        "Write-Output '===='; net user guest | select-string -pattern 'Account active'")

                elif "Administrator account status" in description:
                    check_account_args.append(
                        "Write-Output '===='; net user administrator | select-string -pattern 'Account active'")

                elif "Rename administrator account" in description:
                    check_account_args.append(
                        "Write-Output '===='; net user administrator | select-string -pattern 'User name'")

                elif "Rename guest account" in description:
                    check_account_args.append(
                        "Write-Output '===='; net user guest | select-string -pattern 'User name'")
                else:
                    check_account_args.append("Write-Output '====';")

            ps_args_dict[key] = ';'.join(check_account_args)

        elif key == "BANNER_CHECK":
            banner_check_args = []
            banner_check_args_list = []

            for idx, val in enumerate(checklist_values):
                # if val == 1:
                # generate command list for getting regristry value
                reg_key = str(reg_key_values[idx])
                reg_item = str(reg_item_values[idx])

                if reg_key.startswith("HKLM"):
                    reg_key = reg_key.replace("HKLM", "HKLM:")
                elif reg_key.startswith("HKU"):
                    reg_key = reg_key.replace("HKU", "HKU:")

                arg = f"Write-Output '====';Get-ItemPropertyValue -Path '{reg_key}' -Name '{reg_item}'"
                banner_check_args.append(arg)

                # if len(banner_check_args) == 50:
                #     banner_check_args_list.append(
                #         ';'.join(banner_check_args))
                #     banner_check_args = []

            banner_check_args_list.append(';'.join(banner_check_args))

            ps_args_dict[key] = ';'.join(banner_check_args)

        elif key == "ANONYMOUS_SID_SETTING":

            anonymous_sid_args = []
            anonymous_sid_args_list = []

            for idx, val in enumerate(checklist_values):
                # if val == 1:

                # generate command list for getting password policy value
                description = str(description_values[idx])

                if "Allow anonymous SID/Name translation" in description:
                    subcategory = 'LSAAnonymousNameLookup ='

                arg = f"Write-Output '===='; Get-Content -Path C:\\temp\\secpol.cfg | Select-String -Pattern '{subcategory}'"

                anonymous_sid_args.append(arg)

                # if len(anonymous_sid_args) == 50:
                #     anonymous_sid_args_list.append(
                #         ';'.join(anonymous_sid_args))
                #     anonymous_sid_args = []

            anonymous_sid_args_list.append(';'.join(anonymous_sid_args))

            ps_args_dict[key] = ';'.join(anonymous_sid_args)

            continue

        elif key == "AUDIT_POLICY_SUBCATEGORY":
            audit_policy_args = []
            audit_policy_args_list = []

            for idx, val in enumerate(checklist_values):

                # if val == 1:
                subcategory = str(subcategory_values[idx])

                arg = f"Write-Output '===='; auditpol /get /subcategory:'{subcategory}' | select-string -pattern '{subcategory}'"

                audit_policy_args.append(arg)

                # if len(audit_policy_args) == 50:
                #     audit_policy_args_list.append(
                #         ';'.join(audit_policy_args))
                #     audit_policy_args = []

            audit_policy_args_list.append(';'.join(audit_policy_args))

            ps_args_dict[key] = ';'.join(audit_policy_args)

        elif key == "REG_CHECK":
            reg_check_args = []
            reg_check_args_list = []

            for idx, val in enumerate(checklist_values):

                # if val == 1:

                reg_key = reg_key_values[idx]
                reg_item = reg_item_values[idx]

                if reg_key.startswith("HKLM"):
                    reg_key = reg_key.replace("HKLM", "HKLM:")
                elif reg_key.startswith("HKU"):
                    reg_key = reg_key.replace("HKU", "HKU:")

                arg = f"Write-Output '====';Get-ItemPropertyValue -Path '{reg_key}' -Name '{reg_item}'"

                reg_check_args.append(arg)

            reg_check_args_list.append(';'.join(reg_check_args))

            ps_args_dict[key] = ';'.join(reg_check_args)

        elif key == "WMI_POLICY":
            wmi_policy_args = []
            wmi_policy_args_list = []

            for idx, val in enumerate(checklist_values):

                # if val == 1:

                arg = "Write-Output '====';(Get-WmiObject -Class Win32_ComputerSystem).DomainRole"

                wmi_policy_args.append(arg)

            wmi_policy_args_list.append(';'.join(wmi_policy_args))

            ps_args_dict[key] = ';'.join(wmi_policy_args)

        else:
            continue

    return ps_args_dict


def read_file(fname):
    data_dict = {
        "PASSWORD_POLICY": [],
        "REGISTRY_SETTING": [],
        "LOCKOUT_POLICY": [],
        "USER_RIGHTS_POLICY": [],
        "CHECK_ACCOUNT": [],
        "BANNER_CHECK": [],
        "ANONYMOUS_SID_SETTING": [],
        "AUDIT_POLICY_SUBCATEGORY": [],
        "REG_CHECK": [],
        "WMI_POLICY": []
    }

    xl = pd.ExcelFile(fname)
    # df = xl.parse(sheet_name=0)

    for type in data_dict:
        try:
            data_dict[type] = xl.parse(sheet_name=type)
        except ValueError as e:
            logging.error(f"{type} not found")
            logger.error('Value not found: %s', e)

    return data_dict


def save_file(out_fname, data_dict_list):

    if data_dict_list == []:
        return

    all_frames = []

    for data_dict in data_dict_list:
        frames = []

        for df in data_dict.values():
            frames.append(df)
        result_rowwise = pd.concat(frames)
        all_frames.append(result_rowwise.reset_index(drop=True))

    result = pd.concat(all_frames, axis=1)

    result = result.loc[:, ~result.columns.duplicated()]

    column_names = result.columns.tolist()

    value_n_result = column_names[10:]
    ip_list = []
    name_list = []

    for i in range(len(value_n_result)):
        if i % 2 == 0:
            ip = value_n_result[i].split('|')[0].strip()
            ip_list.append(ip)
            name_list.append('Actual Value')
        else:
            ip_list.append('')
            name_list.append('Result')

    new_data = ['Checklist', 'Type', 'Index', 'Description', 'Reg Key', 'Reg Item', 'Reg Option', 'Audit Policy Subcategory',
                'Right type', 'Value Data'] + name_list
    result.columns = ['Checklist', 'Type', 'Index', 'Description', 'Reg Key', 'Reg Item', 'Reg Option', 'Audit Policy Subcategory',
                      'Right type', 'Value Data'] + ip_list

    new_df = pd.DataFrame(
        [new_data + [''] * (result.shape[1] - len(new_data))], columns=result.columns)
    result = pd.concat([new_df, result]).reset_index(drop=True)

    # Save DataFrame to a new Excel file
    result.to_excel(out_fname, index=False)

    # Load the workbook and select the sheet
    wb = load_workbook(out_fname)
    ws = wb.active

    # Merge the appropriate cells in the new first row
    for col in range(1, 11):  # adjust these values as needed
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=2, end_column=col)

    for ip_col in range(11, len(result.columns)):  # adjust these values as needed
        if ip_col % 2 != 0:
            ws.merge_cells(start_row=1, start_column=ip_col,
                           end_row=1, end_column=ip_col+1)
        else:
            continue

    # Save the workbook
    wb.save(out_fname)
    print((f"Result saved into {out_fname}"))

    logger.info(f"Result saved into {out_fname}")
    logging.info(f"Result saved into {out_fname}")


def configurations(config_fname):

    config_file = pd.ExcelFile(config_fname)

    configs = config_file.parse(sheet_name=0)

    ip_list = []

    ips = configs['IP Address'].values
    usernames = configs['Username'].values
    passwords = configs['Password'].values
    versions = configs['Windows Version'].values

    for idx, val in enumerate(ips):
        ip = [ips[idx], usernames[idx],
              passwords[idx], versions[idx]]

        ip_list.append(ip)

        # verify os
        # os = getOS(ip)
        # if os == versions[idx]:
        #     ip_list.append(ip)
        # else:
        #     print(
        #         f"IP: {ip[0]} | Invalid version | Expected: {ip[3]} | Actual: {os}")

    # return ip_dict
    return ip_list


if __name__ == '__main__':

    my_parser = argparse.ArgumentParser(
        description='A Customizable Multiprocessing Remote Security Audit Program')

    # Add the arguments
    my_parser.add_argument('--audit',
                           type=str,
                           required=True,
                           help='The audit file')

    my_parser.add_argument('--output',
                           type=str,
                           required=False,
                           help='The output powershell file')

    # Execute parse_args()
    args = my_parser.parse_args()

    print('Aduit file:', args.audit)
    print('Output file:', args.output)

    # fname = "src\Audit\CIS_MS_Windows_11_Enterprise_Level_1_v1.0.0.xlsx"

    fname = args.audit
    data_dict = read_file(fname)
    ps_args_dict = gen_ps_args(data_dict)

    if args.output:
        script_name = args.output
    else:

        script_name = 'out\\script\\' + \
            fname.split("\\")[-1].replace("xlsx", "ps1")

    with open(script_name, 'w') as f:
        for key in ps_args_dict:
            for cmd in ps_args_dict[key]:
                f.write(cmd)
            f.write(";")
            # print(cmd)

    print("Done! File saved: %s", script_name)
