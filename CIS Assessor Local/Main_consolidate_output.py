import pandas as pd
import time
import logging
import argparse


from multiprocessing import Pool, Manager
from openpyxl import load_workbook

from getPwdPolicy import compare_pwd_policy_local
from getRegValue import compare_reg_value_local
from getLockoutPolicy import compare_lockout_policy_local
from getUserRights import compare_user_rights_local
from getCheckAccount import compare_check_account_local
from getBannerCheck import compare_banner_check_local
from getAnonySID import compare_anonymous_sid_local
from getAuditPolicy import compare_audit_policy_local
from getRegCheck import compare_reg_check_local
from getWMIPolicy import compare_wmi_policy_local


# set up logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

handler = logging.FileHandler('mylog.log', mode='w')
handler.setLevel(logging.DEBUG)

formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

logger.addHandler(handler)


def get_actual_values(data_dict):

    new_dict = {}

    for key, args_list in data_dict.items():

        try:

            if key == "PASSWORD_POLICY":
                new_df = compare_pwd_policy_local(data_dict)
            elif key == "REGISTRY_SETTING":
                new_df = compare_reg_value_local(data_dict)
            elif key == "LOCKOUT_POLICY":
                new_df = compare_lockout_policy_local(data_dict)
            elif key == "USER_RIGHTS_POLICY":
                new_df = compare_user_rights_local(data_dict)
            elif key == "CHECK_ACCOUNT":
                new_df = compare_check_account_local(data_dict)
            elif key == "BANNER_CHECK":
                new_df = compare_banner_check_local(data_dict)
            elif key == "ANONYMOUS_SID_SETTING":
                new_df = compare_anonymous_sid_local(data_dict)
            elif key == "AUDIT_POLICY_SUBCATEGORY":
                new_df = compare_audit_policy_local(data_dict)
            elif key == "REG_CHECK":
                new_df = compare_reg_check_local(data_dict)
            elif key == "WMI_POLICY":
                new_df = compare_wmi_policy_local(data_dict)

            new_dict[key] = new_df
        except Exception as e:
            logger.error('Failed to get actual value: %s', e)
            logging.debug('Failed to get actual value: %s', e)

        # new_dict[key] = new_df

    return new_dict


def read_file(fname):
    data_dict = {
        "PASSWORD_POLICY": [],
        "REGISTRY_SETTING": [],
        "LOCKOUT_POLICY": [],
        "USER_RIGHTS_POLICY": [],
        "CHECK_ACCOUNT": [],
        "BANNER_CHECK": [],
        "ANONYMOUS_SID_SETTING": [],
        "AUDIT_POLICY_SUBCATEGORY": [],
        "REG_CHECK": [],
        "WMI_POLICY": []
    }

    xl = pd.ExcelFile(fname)
    # df = xl.parse(sheet_name=0)

    def remove_illegal_chars(val):
        if isinstance(val, str):
            # Remove control characters
            val = ''.join(ch for ch in val if ch.isprintable())
        return val

    for ptype in data_dict:
        try:
            df0 = xl.parse(sheet_name=ptype)
            data_dict[ptype] = df0.applymap(remove_illegal_chars)
        except ValueError as e:
            logging.error(f"{ptype} not found")
            logger.error('Value not found: %s', e)

    return data_dict


def save_file(out_fname, data_dict_list, ip_addr):

    if data_dict_list == []:
        return

    all_frames = []

    for data_dict in data_dict_list:
        frames = []

        for df in data_dict.values():
            frames.append(df)
        result_rowwise = pd.concat(frames)
        all_frames.append(result_rowwise.reset_index(drop=True))

    result = pd.concat(all_frames, axis=1)

    result = result.loc[:, ~result.columns.duplicated()]

    column_names = result.columns.tolist()

    value_n_result = column_names[10:]
    ip_list = [ip_addr, '']
    name_list = []

    for i in range(len(value_n_result)):
        if i % 2 == 0:
            # ip = value_n_result[i].split('|')[0].strip()
            # ip_list.append(ip)
            name_list.append('Actual Value')
        else:
            # ip_list.append('')
            name_list.append('Result')

    new_data = ['Checklist', 'Type', 'Index', 'Description', 'Reg Key', 'Reg Item', 'Reg Option', 'Audit Policy Subcategory',
                'Right type', 'Value Data'] + name_list
    result.columns = ['Checklist', 'Type', 'Index', 'Description', 'Reg Key', 'Reg Item', 'Reg Option', 'Audit Policy Subcategory',
                      'Right type', 'Value Data'] + ip_list

    new_df = pd.DataFrame(
        [new_data + [''] * (result.shape[1] - len(new_data))], columns=result.columns)
    result = pd.concat([new_df, result]).reset_index(drop=True)

    def remove_illegal_chars(val):
        if isinstance(val, str):
            # Remove control characters
            val = ''.join(ch for ch in val if ch.isprintable())
        return val

    # Apply the function to each string column in the DataFrame
    result = result.applymap(remove_illegal_chars)

    # Save DataFrame to a new Excel file
    result.to_excel(out_fname, index=False)

    # Load the workbook and select the sheet
    wb = load_workbook(out_fname)
    ws = wb.active

    # Merge the appropriate cells in the new first row
    for col in range(1, 11):  # adjust these values as needed
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=2, end_column=col)

    for ip_col in range(11, len(result.columns)):  # adjust these values as needed
        if ip_col % 2 != 0:
            ws.merge_cells(start_row=1, start_column=ip_col,
                           end_row=1, end_column=ip_col+1)
        else:
            continue

    # Save the workbook
    wb.save(out_fname)
    print((f"Result saved into {out_fname}"))

    logger.info(f"Result saved into {out_fname}")
    logging.info(f"Result saved into {out_fname}")


def configurations(config_fname):

    config_file = pd.ExcelFile(config_fname)

    configs = config_file.parse(sheet_name=0)

    ip_list = []

    ips = configs['IP Address'].values
    usernames = configs['Username'].values
    passwords = configs['Password'].values
    versions = configs['Windows Version'].values

    for idx, val in enumerate(ips):
        ip = [ips[idx], usernames[idx],
              passwords[idx], versions[idx]]

        ip_list.append(ip)

    # return ip_dict
    return ip_list


if __name__ == '__main__':

    my_parser = argparse.ArgumentParser(
        description='A Customizable Multiprocessing Remote Security Audit Program')

    # Add the arguments
    my_parser.add_argument('--result',
                           type=str,
                           required=True,
                           help='The result file')

    my_parser.add_argument('--output',
                           type=str,
                           required=True,
                           help='The output file')

    my_parser.add_argument('--audit',
                           type=str,
                           required=True,
                           help='The audit file')

    # Execute parse_args()
    args = my_parser.parse_args()

    print('Result file:', args.result)
    print('Output file:', args.output)
    print('Audit file:', args.audit)

    start_t0 = time.time()

    # result_fname = "output_win10.txt"
    result_fname = args.result

    output_list = []
    with open(result_fname, 'r', encoding='utf-16') as file:
        lines = file.readlines()

    single_line = ' '.join(line.strip() for line in lines)

    # actual value list
    output_list = single_line.strip().split("====")
    output_list.pop(0)

    print(len(output_list))

    # audit_fname = "src\Audit\CIS_MS_Windows_10_Enterprise_Level_1_v2.0.0.xlsx"
    audit_fname = args.audit

    data_dict = read_file(audit_fname)

    # add actual value to the audit file
    head = 0
    for key in data_dict:
        length = len(data_dict[key])
        # print(key, len(output_list[head: (head+length)]))
        # print()
        # print(output_list[head: (head+length)])
        # print()
        data_dict[key]['Actual Value'] = output_list[head: (head+length)]
        head += length

    # print(data_dict)

    new_dict = get_actual_values(data_dict)
    results = []
    results.append(new_dict)

    ip_addr = "IP"
    # # write output file
    save_file(args.output, results, ip_addr)

    # start_t = time.time()

    # end_t0 = time.time()

    # running_t0 = end_t0 - start_t0
    # logger.info(f"The script ran for {running_t0} seconds")
    # logging.info(f"The script ran for {running_t0} seconds")
