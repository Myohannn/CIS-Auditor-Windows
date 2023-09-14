import pandas as pd
import time
import logging
import argparse
import sys

from openpyxl import load_workbook

from utilities.getPwdPolicy import compare_pwd_policy_local
from utilities.getRegValue import compare_reg_value_local
from utilities.getLockoutPolicy import compare_lockout_policy_local
from utilities.getUserRights import compare_user_rights_local
from utilities.getCheckAccount import compare_check_account_local
from utilities.getBannerCheck import compare_banner_check_local
from utilities.getAnonySID import compare_anonymous_sid_local
from utilities.getAuditPolicy import compare_audit_policy_local
from utilities.getRegCheck import compare_reg_check_local
from utilities.getWMIPolicy import compare_wmi_policy_local


def get_actual_values(data_dict: dict) -> dict:
    '''
    This function takes a dictionary of data as input. For each type of audit, 
    it calls the appropriate function to compare the actual and expected values. 
    The function then returns a new dictionary with the comparison results.

    :param data_dict: 
        A dictionary with keys representing different audit types and values as 
        DataFrames containing the audit data.
    :return: 
        A new dictionary with the same keys as data_dict but the values replaced 
        with DataFrames that include the results of the comparison between the 
        actual and expected values.
    '''

    new_dict = {}

    for key in data_dict.keys():

        try:

            if key == "PASSWORD_POLICY":
                new_df = compare_pwd_policy_local(data_dict)
            elif key == "REGISTRY_SETTING":
                new_df = compare_reg_value_local(data_dict)
            elif key == "LOCKOUT_POLICY":
                new_df = compare_lockout_policy_local(data_dict)
            elif key == "USER_RIGHTS_POLICY":
                new_df = compare_user_rights_local(data_dict)
            elif key == "CHECK_ACCOUNT":
                new_df = compare_check_account_local(data_dict)
            elif key == "BANNER_CHECK":
                new_df = compare_banner_check_local(data_dict)
            elif key == "ANONYMOUS_SID_SETTING":
                new_df = compare_anonymous_sid_local(data_dict)
            elif key == "AUDIT_POLICY_SUBCATEGORY":
                new_df = compare_audit_policy_local(data_dict)
            elif key == "REG_CHECK":
                new_df = compare_reg_check_local(data_dict)
            elif key == "WMI_POLICY":
                new_df = compare_wmi_policy_local(data_dict)

            new_dict[key] = new_df
        except Exception as e:
            print('Failed to get actual value:', e)

    return new_dict


def read_file(fname: str) -> dict:
    '''
    This function reads the provided audit file and returns a dictionary where 
    the keys are different audit types and the values are corresponding audit 
    data in DataFrame format.

    :param fname: 
        A string representing the filename of the audit file.
    :return: 
        A dictionary with keys representing different audit types and values as 
        DataFrames containing the audit data.
    '''

    data_dict = {
        "PASSWORD_POLICY": [],
        "REGISTRY_SETTING": [],
        "LOCKOUT_POLICY": [],
        "USER_RIGHTS_POLICY": [],
        "CHECK_ACCOUNT": [],
        "BANNER_CHECK": [],
        "ANONYMOUS_SID_SETTING": [],
        "AUDIT_POLICY_SUBCATEGORY": [],
        "REG_CHECK": [],
        "WMI_POLICY": []
    }

    xl = pd.ExcelFile(fname)
    # df = xl.parse(sheet_name=0)

    for ptype in data_dict:
        try:
            df0 = xl.parse(sheet_name=ptype)
            data_dict[ptype] = df0.applymap(remove_illegal_chars)
        except ValueError as e:
            logging.error(f"{ptype} not found")

    return data_dict


def remove_illegal_chars(val: str) -> str:
    '''
    This function checks if the given value is a string, and if so, removes any 
    non-printable characters.

    :param val: 
        A string potentially containing non-printable characters.
    :return: 
        The same string but with any non-printable characters removed.
    '''

    if isinstance(val, str):
        # Remove control characters
        val = ''.join(ch for ch in val if ch.isprintable())
    return val


def save_file(out_fname: str, data_dict_list: list, ip_addr: str) -> None:
    ''' 
    This function processes the comparison results and saves them into an Excel file.

    :param out_fname: 
        A string representing the filename of the output file.
    :param data_dict_list: 
        A list of dictionaries containing comparison results.
    :param ip_addr: 
        A string representing an IP address.
    :return: 
        None
    '''

    if data_dict_list == []:
        return

    all_frames = []

    for data_dict in data_dict_list:
        frames = []

        for df in data_dict.values():
            frames.append(df)
        result_rowwise = pd.concat(frames)
        all_frames.append(result_rowwise.reset_index(drop=True))

    result = pd.concat(all_frames, axis=1)

    result = result.loc[:, ~result.columns.duplicated()]

    column_names = result.columns.tolist()

    value_n_result = column_names[11:]
    ip_list = [ip_addr, '']
    name_list = []

    for i in range(len(value_n_result)):
        if i % 2 == 0:
            # ip = value_n_result[i].split('|')[0].strip()
            # ip_list.append(ip)
            name_list.append('Actual Value')
        else:
            # ip_list.append('')
            name_list.append('Result')

    new_data = ['Checklist', 'Type', 'Index', 'Description', 'Solution', 'Reg Key', 'Reg Item', 'Reg Option', 'Audit Policy Subcategory',
                'Right type', 'Value Data'] + name_list
    result.columns = ['Checklist', 'Type', 'Index', 'Description', 'Solution', 'Reg Key', 'Reg Item', 'Reg Option', 'Audit Policy Subcategory',
                      'Right type', 'Value Data'] + ip_list

    new_df = pd.DataFrame(
        [new_data + [''] * (result.shape[1] - len(new_data))], columns=result.columns)
    result = pd.concat([new_df, result]).reset_index(drop=True)

    # Apply the function to each string column in the DataFrame
    result = result.applymap(remove_illegal_chars)

    # Save DataFrame to a new Excel file
    result.to_excel(out_fname, index=False)

    # Load the workbook and select the sheet
    wb = load_workbook(out_fname)
    ws = wb.active

    # Merge the appropriate cells in the new first row
    for col in range(1, 12):  # adjust these values as needed
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=2, end_column=col)

    for ip_col in range(12, len(result.columns)):  # adjust these values as needed
        if ip_col % 2 == 0:
            ws.merge_cells(start_row=1, start_column=ip_col,
                           end_row=1, end_column=ip_col+1)
        else:
            continue

    # Save the workbook
    wb.save(out_fname)
    print((f"Result saved into {out_fname}"))

    logging.info(f"Result saved into {out_fname}")


'''
This is the main function that gets executed when the script runs. It parses 
command-line arguments, reads the audit file, compares actual and expected 
values for each audit type, and finally saves the comparison results into an 
output file.
'''
if __name__ == '__main__':

    my_parser = argparse.ArgumentParser(
        description="This is a script for processing audit results.")

    # Add the arguments
    my_parser.add_argument('-ps_result',
                           type=str,
                           required=True,
                           help='(REQUIRED) The path of the result file generated after running the PowerShell script. This should be a .txt file.')

    my_parser.add_argument('-output',
                           type=str,
                           required=True,
                           help='(REQUIRED) The path where the output result file should be saved. This should be a .txt file.')

    my_parser.add_argument('-audit',
                           type=str,
                           required=True,
                           help='(REQUIRED) The path of the parsed audit file that contains the audit results you want to process. This should be a .xlsx file')

    # Execute parse_args()
    try:
        args = my_parser.parse_args()
    except SystemExit:
        my_parser.print_help()
        sys.exit(1)

    print('PowerShell result file:', args.ps_result)
    print('Output file:', args.output)
    print('Audit file:', args.audit)

    start_t0 = time.time()

    # result_fname = "output_win10.txt"
    result_fname = args.ps_result

    output_list = []
    with open(result_fname, 'r', encoding='utf-16') as file:
        lines = file.readlines()

    single_line = ' '.join(line.strip() for line in lines)

    # actual value list
    output_list = single_line.strip().split("====")
    output_list.pop(0)
    ip_addr = output_list[0]
    output_list.pop(0)

    # audit_fname = "src\Audit\CIS_MS_Windows_10_Enterprise_Level_1_v2.0.0.xlsx"
    audit_fname = args.audit

    data_dict = read_file(audit_fname)

    # add actual value to the audit file
    head = 0
    for key in data_dict:
        length = len(data_dict[key])
        data_dict[key]['Actual Value'] = output_list[head: (head+length)]
        head += length

    # print(data_dict)

    new_dict = get_actual_values(data_dict)
    results = []
    results.append(new_dict)

    # # write output file
    save_file(args.output, results, ip_addr)
